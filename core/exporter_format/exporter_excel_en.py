from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def build_excel_en(violations, page_name, page_url, stats, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Accessibility Report"

    # 定义样式常量
    thin = Side(border_style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2c3e50")
    critical_fill = PatternFill("solid", fgColor="e74c3c")  # 红色(A级)
    serious_fill = PatternFill("solid", fgColor="f39c12")   # 橙色(AA级)
    normal_font = Font(size=10)
    bold_font = Font(size=10, bold=True)
    title_font = Font(size=14, bold=True)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    disclaimer_font = Font(size=9, color="7f8c8d")

    # ----------------------
    # 顶部信息（保留原有结构+新增URL/时间/工具/合规标准）
    # ----------------------
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Website Accessibility Test Report"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill("solid", fgColor="f8f9fa")
    title_cell.border = border

    # 基础信息（保留原有Page/URL，新增Test Time/Tool/Compliance Standard）
    ws['A3'] = "URL"
    ws['B3'] = page_url
    ws['A4'] = "Test Time"
    ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 动态时间
    ws['C4'] = "Testing Tool"
    ws['D4'] = "Axe DevTools"
    ws['E4'] = "Compliance Standard"
    ws['F4'] = "wcag2a | wcag2aa | wcag21a | wcag21aa | section508 | EN-301-549"

    # 原有统计信息结构完全保留（仅优化样式）
    ws['A5'] = "Total Issues"
    ws['B5'] = stats['total']
    ws['C5'] = "Critical"
    ws['D5'] = stats['critical']
    ws['E5'] = "Serious"
    ws['F5'] = stats['serious']
    ws['A6'] = "Moderate"
    ws['B6'] = stats['moderate']
    ws['C6'] = "Minor"
    ws['D6'] = stats['minor']
    ws['E6'] = "Compliance Rate"
    ws['F6'] = f"{stats['compliance_rate']}%"

    # 顶部信息样式统一优化
    for r in range(2, 7):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = bold_font if c % 2 == 1 else normal_font  # 标签加粗
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.fill = PatternFill("solid", fgColor="e9ecef") if c % 2 == 1 else PatternFill("solid", fgColor="f8f9fa")

    # ----------------------
    # 表头（新增Fix Recommendation列，保留原有所有列）
    # ----------------------
    header_row = 8
    # 原有表头 + 新增Fix Recommendation列
    headers = ["No.", "ID", "Severity", "WCAG", "Element Locator", "Description", "Fix Recommendation"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # ----------------------
    # 填充违规数据（完全保留原有判断逻辑+新增Fix Recommendation）
    # ----------------------
    row = header_row + 1
    no = 1

    for v in violations:
        tags = v.get('tags', [])
        # 【完全保留原有WCAG等级判断逻辑】
        wcag_level = []
        for t in v.get('tags', []):
            if t in ["wcag2a","wcag21a"]:
                wcag_level.append("A")
            elif t in ["wcag2aa","wcag21aa"]:
                wcag_level.append("AA")
            elif t in ["wcag2aaa","wcag21aaa"]:
                wcag_level.append("AAA")
        wcag = "/".join(sorted(set(wcag_level))) if wcag_level else "Best Practice"

        # 【完全保留原有WCAG条款提取逻辑】
        wcag_clauses = []
        for t in tags:
            if t.startswith("wcag") and len(t) >= 5 and t[4:].replace(".","").isdigit():
                num_part = t[4:]
                clause = ".".join(list(num_part))
                wcag_clauses.append(clause)
        clause_str = " / ".join(sorted(list(set(wcag_clauses)))) if wcag_clauses else ""
        if clause_str:
            wcag = f"WCAG {clause_str} ({wcag})"
        else:
            wcag = wcag

        # 【完全保留原有元素定位器提取逻辑】
        selectors = []
        for node in v.get('nodes', []):
            selectors.extend(node.get('target', []))
        selectors = list(set(selectors)) or ["N/A"]
        count = len(selectors)

        # 【新增】修复建议（优先取help字段，无则给默认值）
        fix_recommendation = v.get('help', "Refer to WCAG guidelines for remediation steps")

        # 【完全保留原有序号+元素定位器写入逻辑】
        for idx, sel in enumerate(selectors):
            # 序号
            no_cell = ws.cell(row=row+idx, column=1, value=no)
            no_cell.alignment = Alignment(horizontal='center', vertical='center')
            no_cell.font = normal_font
            no_cell.border = border
            # 元素定位器
            sel_cell = ws.cell(row=row+idx, column=5, value=sel)
            sel_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            sel_cell.font = normal_font
            sel_cell.border = border
            no += 1

        # 【保留原有合并逻辑+适配新增列】
        if count > 1:
            ws.merge_cells(f'B{row}:B{row+count-1}')
            ws.merge_cells(f'C{row}:C{row+count-1}')
            ws.merge_cells(f'D{row}:D{row+count-1}')
            ws.merge_cells(f'F{row}:F{row+count-1}')
            ws.merge_cells(f'G{row}:G{row+count-1}')  # 新增列合并

        # 【保留原有合并内容写入逻辑】
        # ID
        id_cell = ws.cell(row=row, column=2, value=v['id'])
        id_cell.font = bold_font
        id_cell.alignment = Alignment(horizontal='left', vertical='center')
        id_cell.border = border
        # Severity
        severity_cell = ws.cell(row=row, column=3, value=v['impact'])
        severity_cell.font = normal_font
        severity_cell.alignment = Alignment(horizontal='left', vertical='center')
        severity_cell.border = border
        # WCAG（新增颜色标记）
        wcag_cell = ws.cell(row=row, column=4, value=wcag)
        wcag_cell.font = normal_font
        wcag_cell.alignment = Alignment(horizontal='left', vertical='center')
        wcag_cell.border = border
        # 【新增】WCAG等级颜色标记
        if "A" in wcag and "AA" not in wcag:
            wcag_cell.fill = critical_fill
            wcag_cell.font = Font(color="FFFFFF", size=10)  # 白色文字
        elif "AA" in wcag:
            wcag_cell.fill = serious_fill
            wcag_cell.font = Font(color="FFFFFF", size=10)  # 白色文字
        # Description
        desc_cell = ws.cell(row=row, column=6, value=v['description'])
        desc_cell.font = normal_font
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        desc_cell.border = border
        # 【新增】Fix Recommendation
        fix_cell = ws.cell(row=row, column=7, value=fix_recommendation)
        fix_cell.font = normal_font
        fix_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        fix_cell.border = border

        # 【保留原有边框逻辑+适配新增列】
        for r in range(row, row+count):
            for c in range(1, 8):  # 列数从6→7（新增Fix Recommendation）
                cell = ws.cell(row=r, column=c)
                if cell.font is None:
                    cell.font = normal_font
                if cell.border is None:
                    cell.border = border

        row += count

    # ----------------------
    # 【新增】底部免责声明
    # ----------------------
    disclaimer_row = row + 2
    ws.merge_cells(f'A{disclaimer_row}:G{disclaimer_row+3}')
    disclaimer_text = (
        "1. Fix recommendations are for technical guidance only; actual remediation requires adjustment based on project code.\n"
        "2. Report generation time: 2026-04-19. Test results are only valid for the version at that time.\n"
        "3. This report is generated based on Axe DevTools test results, for reference only. No legal liability is assumed for any consequences arising from the use of this report."
    )
    disclaimer_cell = ws.cell(row=disclaimer_row, column=1, value=disclaimer_text)
    disclaimer_cell.font = disclaimer_font
    disclaimer_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    disclaimer_cell.border = border
    disclaimer_cell.fill = PatternFill("solid", fgColor="f8f9fa")

    # ----------------------
    # 列宽优化（适配新增列）
    # ----------------------
    ws.column_dimensions['A'].width = 8    # No.
    ws.column_dimensions['B'].width = 20   # ID
    ws.column_dimensions['C'].width = 12   # Severity
    ws.column_dimensions['D'].width = 20   # WCAG
    ws.column_dimensions['E'].width = 50   # Element Locator
    ws.column_dimensions['F'].width = 60   # Description
    ws.column_dimensions['G'].width = 50   # Fix Recommendation

    # 保存文件
    wb.save(path)