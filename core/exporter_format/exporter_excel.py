from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build_excel(violations, page_name, page_url, stats, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "无障碍检测报告"

    thin = Side(border_style="thin", color="cccccc")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----------------------
    # 顶部信息（全部保留）
    # ----------------------
    ws.merge_cells('A1:F1')
    ws['A1'] = "网站无障碍检测报告"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = "页面名称"
    ws['B2'] = page_name
    ws['A3'] = "页面地址"
    ws['B3'] = page_url
    ws['A4'] = "总缺陷数"
    ws['B4'] = stats['total']
    ws['C4'] = "严重"
    ws['D4'] = stats['critical']
    ws['E4'] = "较严重"
    ws['F4'] = stats['serious']
    ws['A5'] = "一般"
    ws['B5'] = stats['moderate']
    ws['C5'] = "轻微"
    ws['D5'] = stats['minor']
    ws['E5'] = "合规率"
    ws['F5'] = f"{stats['compliance_rate']}%"

    # 明细表头
    header_row = 7
    headers = ["序号", "缺陷ID", "等级", "WCAG", "元素定位", "问题描述"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="2c3e50")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = header_row + 1
    no = 1

    for v in violations:
        # WCAG
        wcag_level = []
        for t in v.get('tags', []):
            if t in ["wcag2a","wcag21a"]:
                wcag_level.append("A")
            elif t in ["wcag2aa","wcag21aa"]:
                wcag_level.append("AA")
            elif t in ["wcag2aaa","wcag21aaa"]:
                wcag_level.append("AAA")
        wcag = "/".join(sorted(set(wcag_level))) if wcag_level else "Best Practice"

        # 元素列表
        selectors = []
        for node in v.get('nodes', []):
            selectors.extend(node.get('target', []))
        selectors = list(set(selectors)) or ["无"]
        count = len(selectors)

        # 序号 + 元素定位
        for idx, sel in enumerate(selectors):
            ws.cell(row=row+idx, column=1, value=no).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row+idx, column=5, value=sel)
            ws.cell(row=row+idx, column=5).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row+idx, column=5).border = border
            no += 1

        # 合并重复列
        if count > 1:
            ws.merge_cells(f'B{row}:B{row+count-1}')
            ws.merge_cells(f'C{row}:C{row+count-1}')
            ws.merge_cells(f'D{row}:D{row+count-1}')
            ws.merge_cells(f'F{row}:F{row+count-1}')

        # 写入合并内容
        ws.cell(row=row, column=2, value=v['id']).font = Font(bold=True)
        ws.cell(row=row, column=3, value=v['impact'])
        ws.cell(row=row, column=4, value=wcag)
        ws.cell(row=row, column=6, value=v['description'])

        # 对齐 & 边框
        for col in [2,3,4,6]:
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 全表格边框
        for r in range(row, row+count):
            for c in range(1,7):
                ws.cell(row=r, column=c).border = border

        row += count

    # 列宽优化
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60

    wb.save(path)